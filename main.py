import pandas as pd
import openpyxl
used_seri_xlsx = 'used.xlsx'  # file excel xlsx
used_seri_sheet = 'Sheet2'  # sheet contain used seri
used_seri_col = 'B'  # column series
used_seri_col_name = 'so_stk'  # column name of used series

used_series_star_end_only = [
    ("AB. 0", 256041, 257500),
    ("AC", 4122001, 4122101),
    ("AB ", 5016501, 5018500)
]
# ---------input số series sổ đã sử dụng từ file excel xuất từ phần mềm
data_series_uesed_file = pd.read_excel(used_seri_xlsx, sheet_name=used_seri_sheet, usecols=used_seri_col)
series_used_file = list(data_series_uesed_file[used_seri_col_name])


class SeriesStk:
    def __init__(self, first2, start, end):
        self.first2 = first2
        self.start = start
        self.end = end


used_seri = []
for i in range(0, len(used_series_star_end_only)):
    used_seri.append(SeriesStk(first2=used_series_star_end_only[i][0] ,start=used_series_star_end_only[i][1], end=used_series_star_end_only[i][2]))


def make_series(start_end):
    series_stk = []
    for seri in start_end:
        amount_seri = seri.end - seri.start +1
        for i in range(amount_seri):
            series_stk.append(seri.first2 + str(seri.start))
            seri.start += 1
    return series_stk


def compare_series(list_series_1, list_series_2):  # xem series 1 trùng series 2 những số nào
    compared_series = []
    for seri in list_series_1:
        if seri not in list_series_2:
            compared_series.append(seri)
    return compared_series


def write_to_excel(sheet_name, series_name):
    workbook_excel = openpyxl.load_workbook("output.xlsx")
    worksheet = workbook_excel[sheet_name]
    last_row = worksheet.max_row
    if last_row > 1:
        rows_to_delete = range(1, last_row)  # Excludes the first row (index 0)
        # Delete the rows in reverse order to avoid shifting issues
        for row_index in sorted(rows_to_delete, reverse=True):
            worksheet.delete_rows(row_index + 1)  # Adding 1 to adjust for 0-indexing
    last_row = worksheet.max_row
    for i, j in enumerate(series_name, start=1):
        worksheet.cell(row=i + last_row, column=1, value=j)
    workbook_excel.save("output.xlsx")
    workbook_excel.close()


seri_used_start_end_all = make_series(used_seri)
so_hong = compare_series(seri_used_start_end_all, series_used_file)


write_to_excel(sheet_name='Sheet1', series_name=so_hong)
